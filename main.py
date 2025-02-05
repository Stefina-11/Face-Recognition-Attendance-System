import os
import cv2
import face_recognition
from datetime import datetime
import openpyxl
import subprocess
import threading  # Import threading

# Initialize camera
video_capture = cv2.VideoCapture(0)
if not video_capture.isOpened():
    raise ValueError("Could not open video device. Check if the camera is properly connected.")

# Define the path for the Excel file
excel_file_path = "Attendance.xlsx"

# Load or create the workbook and sheet
if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
else:
    workbook = openpyxl.Workbook()

sheet = workbook.active
sheet.title = "Attendance"

# Initialize the header if not present
if sheet.max_row == 1:
    sheet.append(["S.No", "RRN", "Name/Email"])  # Initial headers for the attendance sheet

# Define known faces directory and data
KNOWN_FACES_DIR = "D:\\known_faces"
known_face_encodings = []
known_face_names = []
known_face_rrns = []

# Load known faces
def load_known_faces():
    for filename in os.listdir(KNOWN_FACES_DIR):
        if filename.endswith(".jpg") or filename.endswith(".png"):
            image_path = os.path.join(KNOWN_FACES_DIR, filename)
            image = face_recognition.load_image_file(image_path)
            encodings = face_recognition.face_encodings(image)
            if encodings:
                known_face_encodings.append(encodings[0])
                name = filename.split("[")[0].strip()  # Assuming name is part of the filename
                rrn = filename.split("[")[1].split("]")[0].strip()  # Assuming RRN is in brackets
                known_face_names.append(name)
                known_face_rrns.append(rrn)
            else:
                print(f"No face found in {filename}")

load_known_faces()

# Function to mark attendance in Excel
def mark_attendance(rrn, email):
    today_date = datetime.now().strftime("%d/%m/%Y")

    # Check if today's date is already a column header
    date_column = None
    for col in range(4, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == today_date:
            date_column = col
            break
    if date_column is None:
        date_column = sheet.max_column + 1
        sheet.cell(row=1, column=date_column).value = today_date

    # Find or add student row
    student_row = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=3).value == email and sheet.cell(row=row, column=2).value == rrn:
            student_row = row
            break
    if student_row is None:
        student_row = sheet.max_row + 1
        sheet.cell(row=student_row, column=1).value = student_row - 1  # S.No
        sheet.cell(row=student_row, column=2).value = rrn             # RRN
        sheet.cell(row=student_row, column=3).value = email           # Name/Email

    # Mark attendance for today
    sheet.cell(row=student_row, column=date_column).value = "✔"

    # Save workbook
    try:
        workbook.save(excel_file_path)
        print(f"Attendance marked for {email} with RRN {rrn} on {today_date}")
    except Exception as e:
        print(f"Error saving the Excel file: {e}")

# Function to open Excel file in a separate thread
def open_excel():
    try:
        subprocess.Popen(['start', excel_file_path], shell=True)  # Opens the Excel file
        print("Excel file opened successfully.")
    except Exception as e:
        print(f"Error opening Excel file: {e}")

# Process video frames for face recognition and attendance marking
attendance_logged = False  # Flag to check if attendance has been logged
while not attendance_logged:
    ret, frame = video_capture.read()
    if not ret:
        print("Failed to read frame from the camera.")
        break

    # Convert frame to RGB format for face recognition
    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    # Detect face locations and encodings in the current frame
    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name, rrn = "Unknown", None

        if True in matches:
            first_match_index = matches.index(True)
            name = known_face_names[first_match_index]
            rrn = known_face_rrns[first_match_index]

            # Log attendance if not already logged
            logged_names = [sheet.cell(row=row, column=3).value for row in range(2, sheet.max_row + 1)]
            if name not in logged_names:
                mark_attendance(rrn, name)
                attendance_logged = True  # Set the flag to true to exit the loop

                # Open the recorded Excel sheet in a separate thread after logging attendance
                threading.Thread(target=open_excel).start()
                break  # Exit loop after logging attendance

        # Draw a rectangle around the face and display the name
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

    # Display the frame with the attendance system
    cv2.imshow('Attendance System', frame)

    # Check for quit command
    if cv2.waitKey(1) & 0xFF == ord('q'):
        attendance_logged = True  # Exit the loop if 'q' is pressed

# Final cleanup
video_capture.release()
cv2.destroyAllWindows()
